#Max 13 Şubeye Kadar Kabul Eder!!!
from reportlab.platypus import TableStyle,Table
from reportlab.lib import colors,pagesizes
from reportlab.pdfgen import canvas
from tkinter import filedialog
from reportlab.graphics import renderPDF
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import *
import openpyxl,win32com.client,threading
import pandas as pd

def excelfilecommand1():
    global sinav1
    sinav1 = filedialog.askopenfilename()

def excelfilecommand2():
    global sinav2
    sinav2 = filedialog.askopenfilename()
    
def excelfilecommand3():
    global sinav3
    sinav3 = filedialog.askopenfilename()

def outputfoldercommand():
    global outputfolder
    outputfolder = filedialog.askdirectory()

def rlreport():
    book1 = openpyxl.load_workbook(sinav1)
    sheet1 = book1['Sheet2']
    book2 = openpyxl.load_workbook(sinav2)
    sheet2 = book2['Sheet2']

    #İki excelde öğrencileri karşılaştırıp sonra ikisine de eksik olan öğrencileri ekliyor.
    df1 = pd.read_excel(sinav1, 'Sheet2')
    df2 = pd.read_excel(sinav2, 'Sheet2')
    df3  = pd.merge(df1, df2, on="Student ID", how='outer', indicator='Exist')
    df3  = df3.loc[df3['Exist'] != 'both']
    fark = len(df3)
    if fark > 0:
        x = -1
        while x<fark-1:
            x+=1
            if type(df3.iloc[x]['Institution Name_x']) == str:
                z1 = df3.iloc[x]['Institution Name_x']
                z2 = df3.iloc[x]['Institution ID_x']
                z3 = df3.iloc[x]['Class_x']
                z4 = df3.iloc[x]['Family Name_x']
                z5 = df3.iloc[x]['Given Name_x']
                z6 = int(df3.iloc[x]['Student ID'])
                z7 = str(df3.iloc[x]['Date of Birth_x'])[:-9]
                print(z1,z2,z3,z4,z5,z6,z7)
                sheet2.append([z1,z2,z3,z4,z5,z6,z7,"Girmedi","_","_","_","_","_","_","_"])
            book2.save(sinav2)
    df4  = pd.merge(df2, df1, on="Student ID", how='outer', indicator='Exist')
    df4  = df4.loc[df4['Exist'] != 'both']
    fark = len(df4)
    if fark > 0:
        x = -1
        while x<fark-1:
            x+=1
            if type(df4.iloc[x]['Institution Name_x']) == str:
                zz1 = df4.iloc[x]['Institution Name_x']
                zz2 = df4.iloc[x]['Institution ID_x']
                zz3 = df4.iloc[x]['Class_x']
                zz4 = df4.iloc[x]['Family Name_x']
                zz5 = df4.iloc[x]['Given Name_x']
                zz6 = int(df4.iloc[x]['Student ID'])
                zz7 = str(df4.iloc[x]['Date of Birth_x'])[:-9]
                print(zz1,zz2,zz3,zz4,zz5,zz6,zz7)
                sheet1.append([zz1,zz2,zz3,zz4,zz5,zz6,zz7,"Girmedi","_","_","_","_","_","_","_"])
            book1.save(sinav1)

    #Excelde sıralama yapıyor böylece excelin hiç düzenlemesine ihtiyaç kalmıyor.
    sheetlenght1 = len(sheet1['A'])
    excelwin = win32com.client.Dispatch("Excel.Application")

    wb = excelwin.Workbooks.Open(sinav1)
    ws = wb.Worksheets('Sheet2')
    ws.Range('A3:O'+str(sheetlenght1)).Sort(Key1=ws.Range('D1'), Order1=1, Orientation=1)
    ws.Range('A3:O'+str(sheetlenght1)).Sort(Key1=ws.Range('C1'), Order1=1, Orientation=1)
    wb.Save()
    excelwin.Application.Quit()

    classlist = []
    for row in sheet1.rows:
        classlist.append(row[2].value)
    my_dict = {i:classlist.count(i) for i in classlist}
    b = list(set([x for x in classlist if classlist.count(x) > 2]))
    b.sort()
    filigram = './filigram.jpg'
    data1= []
    kk1 = []
    
    def sonuclar():
        satir = 2
        toplamsatir = int(sheetlenght1)-1
        data= [['','Student Name', 'Class', 'Reading\nScore\nCEFR\n'+str(ay1)+'\n'+str(yil1),'', 'Reading\nScore\nCEFR\n'+str(ay2)+'\n'+str(yil2),'', 'Listening\nScore\nCEFR\n'+str(ay1)+'\n'+str(yil1),'','Listening\nScore\nCEFR\n'+str(ay2)+'\n'+str(yil2),'']]
        chartdata = []
        
        while satir<toplamsatir+1:
            satir +=1
            studentnumber = int(sheet1['F'+str(satir)].value)
            studentname = str(sheet1['E'+str(satir)].value)+str(" ")+str(sheet1['D'+str(satir)].value)
            sinif = sheet1['C'+str(satir)].value
            rs1a = sheet1['I'+str(satir)].value
            rs1b = sheet1['J'+str(satir)].value
            ls1a = sheet1['M'+str(satir)].value
            ls1b = sheet1['N'+str(satir)].value
            for row in sheet2.rows:
                for any_cell in row:
                    if any_cell.value == studentnumber:
                            rs2a = sheet2.cell(row=any_cell.row, column=9).value
                            rs2b = sheet2.cell(row=any_cell.row, column=10).value
                            ls2a = sheet2.cell(row=any_cell.row, column=13).value
                            ls2b = sheet2.cell(row=any_cell.row, column=14).value
            
            if b[x1] == sinif:
                rs1b = rs1b.replace('Below A1','*')
                ls1b = ls1b.replace('Below A1','*')
                rs2b = rs2b.replace('Below A1','*')
                ls2b = ls2b.replace('Below A1','*')
                tumveriler = ["",studentname, sinif, rs1a, rs1b,rs2a,rs2b, ls1a,ls1b,ls2a,ls2b]
                chartdata.append(tumveriler)
                data.append(tumveriler)
                table = Table(data)
                style = TableStyle([
                    ('GRID',(1,0),(-1,-1),0.5,colors.grey),
                    ('GRID',(1,1),(-1,-1),0.5,colors.grey),
                    ('ALIGN',(0,0),(-1,0),'CENTER'),
                    ('VALIGN',(0,0),(-1,0),'MIDDLE'),
                    ('TOPPADDING',(0,0),(-1,0),12),
                    ('BOTTOMPADDING',(0,0),(-1,0),12),
                    ('RIGHTPADDING',(1,0),(1,0),38),
                    ('LEFTPADDING',(1,0),(1,0),38),
                    
                    ('ALIGN',(2,0),(-1,-1),'CENTER'),
                    ('VALIGN',(2,0),(-1,1),'MIDDLE'),
                    ('RIGHTPADDING',(3,0),(-1,0),18),
                    ('LEFTPADDING',(3,0),(-1,0),18),
                    
                    ('ALIGN',(0,1),(0,-1),'CENTER'),
                    ('VALIGN',(0,1),(0,-1),'MIDDLE'),
                    ('BACKGROUND',(3,0),(4,0),colors.lightblue),
                    ('BACKGROUND',(5,0),(6,0),colors.lemonchiffon),
                    ('BACKGROUND',(7,0),(8,0),colors.lightblue),
                    ('BACKGROUND',(9,0),(10,0),colors.lemonchiffon),
                    ('SPAN',(3,0),(4,0)),
                    ('SPAN',(5,0),(6,0)),
                    ('SPAN',(7,0),(8,0)),
                    ('SPAN',(9,0),(10,0)),
                    ('FONTSIZE', (0,1), (-1,-1), 8)
                ])
                u1 = 680
                k1 = my_dict[b[x1]]
                
                pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
                pdf.setFont('Helvetica',10)
                pdf.drawString(72,u1,'Class : '+str(b[x1]))
                pdf.drawString(170,u1,'Total Student Number : '+str(k1))
                pdf.setFont('Helvetica',9)
                pdf.drawString(75,100,'CEFR : The Common European Framework of Reference')
                pdf.drawString(75,87,'* : Below A1')
                pdf.drawString(75,74,'- : Do not take the test')
                table.setStyle(style)
                table.wrapOn(pdf, 150, 150)
                table.drawOn(pdf, 60,(u1-95-18*k1))
        kk1.append(k1)
        data1.append(chartdata)
        
    def chart1():
        pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
        pdf.drawCentredString(307.5,490,"Reading Score Comparison 2020-2021")
        pdf.drawCentredString(307.5,290,"Listening Score Comparison 2020-2021")
        pdf.setFont('Helvetica',9)
        pdf.drawString(95, 83,"2020 :")
        pdf.drawString(95, 63,"2021 :")
        
        drawing3 = Drawing(400, 200)
        drawing3.add(Rect(10, 10, 70, 13, fillColor=colors.lightblue))
        x, y = 120, 70 # coordinates (from left bottom)
        renderPDF.draw(drawing3, pdf, x, y, showBoundary=False)
        drawing4 = Drawing(400, 200)
        drawing4.add(Rect(10, 10, 70, 13, fillColor=colors.lemonchiffon))
        x, y = 120, 50 # coordinates (from left bottom)
        renderPDF.draw(drawing4, pdf, x, y, showBoundary=False)

        x2=-1
        r1,r2,r3,r4,r5,r6,r7,r8,r9,r10,r11,r12,r13,r1a,r1b,r1c,r2a,r2b,r2c,r3a,r3b,r3c,r4a,r4b,r4c,r5a,r5b,r5c,r6a,r6b,r6c,r7a,r7b,r7c,r8a,r8b,r8c,r9a,r9b,r9c,r10a,r10b,r10c,r11a,r11b,r11c,r12a,r12b,r12c,r13a,r13b,r13c=[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
        while x2<len(data1)-1:
            x2+=1
            y=-1
            while y<len(data1[x2])-1:
                y+=1
                d11 = data1[x2][y][3]
                d12 = data1[x2][y][5]
                d13 = data1[x2][y][7]
                d14 = data1[x2][y][9]
                if d11 == "*" or d11=="_":
                    d11 = 0
                if d12 == "*" or d12=="_":
                    d12 = 0
                if d13 == "*" or d13=="_":
                    d13 = 0
                if d14 == "*" or d14=="_":
                    d14 = 0
                if d11 > 0:
                    try:
                        if data1[x2][y][2]==b[0]:
                            r1.append(d11)
                        if data1[x2][y][2]==b[1]:
                            r2.append(d11)
                        if data1[x2][y][2]==b[2]:
                            r3.append(d11)
                        if data1[x2][y][2]==b[3]:
                            r4.append(d11)
                        if data1[x2][y][2]==b[4]:
                            r5.append(d11)
                        if data1[x2][y][2]==b[5]:
                            r6.append(d11)
                        if data1[x2][y][2]==b[6]:
                            r7.append(d11)
                        if data1[x2][y][2]==b[7]:
                            r8.append(d11)
                        if data1[x2][y][2]==b[8]:
                            r9.append(d11)
                        if data1[x2][y][2]==b[9]:
                            r10.append(d11)
                        if data1[x2][y][2]==b[10]:
                            r11.append(d11)
                        if data1[x2][y][2]==b[11]:
                            r12.append(d11)
                        if data1[x2][y][2]==b[12]:
                            r13.append(d11)
                    except:
                        pass
                if d12 > 0:
                    try:
                        if data1[x2][y][2]==b[0]:
                            r1a.append(d12)
                        if data1[x2][y][2]==b[1]:
                            r2a.append(d12)
                        if data1[x2][y][2]==b[2]:
                            r3a.append(d12)
                        if data1[x2][y][2]==b[3]:
                            r4a.append(d12)
                        if data1[x2][y][2]==b[4]:
                            r5a.append(d12)
                        if data1[x2][y][2]==b[5]:
                            r6a.append(d12)
                        if data1[x2][y][2]==b[6]:
                            r7a.append(d12)
                        if data1[x2][y][2]==b[7]:
                            r8a.append(d12)
                        if data1[x2][y][2]==b[8]:
                            r9a.append(d12)
                        if data1[x2][y][2]==b[9]:
                            r10a.append(d12)
                        if data1[x2][y][2]==b[10]:
                            r11a.append(d12)
                        if data1[x2][y][2]==b[11]:
                            r12a.append(d12)
                        if data1[x2][y][2]==b[12]:
                            r13a.append(d12)
                    except:
                        pass
                if d13 > 0:
                    try:
                        if data1[x2][y][2]==b[0]:
                            r1b.append(d13)
                        if data1[x2][y][2]==b[1]:
                            r2b.append(d13)
                        if data1[x2][y][2]==b[2]:
                            r3b.append(d13)
                        if data1[x2][y][2]==b[3]:
                            r4b.append(d13)
                        if data1[x2][y][2]==b[4]:
                            r5b.append(d13)
                        if data1[x2][y][2]==b[5]:
                            r6b.append(d13)
                        if data1[x2][y][2]==b[6]:
                            r7b.append(d13)
                        if data1[x2][y][2]==b[7]:
                            r8b.append(d13)
                        if data1[x2][y][2]==b[8]:
                            r9b.append(d13)
                        if data1[x2][y][2]==b[9]:
                            r10b.append(d13)
                        if data1[x2][y][2]==b[10]:
                            r11b.append(d13)
                        if data1[x2][y][2]==b[11]:
                            r12b.append(d13)
                        if data1[x2][y][2]==b[12]:
                            r13b.append(d13)
                    except:
                        pass
                if d14 > 0:
                    try:
                        if data1[x2][y][2]==b[0]:
                            r1c.append(d14)
                        if data1[x2][y][2]==b[1]:
                            r2c.append(d14)
                        if data1[x2][y][2]==b[2]:
                            r3c.append(d14)
                        if data1[x2][y][2]==b[3]:
                            r4c.append(d14)
                        if data1[x2][y][2]==b[4]:
                            r5c.append(d14)
                        if data1[x2][y][2]==b[5]:
                            r6c.append(d14)
                        if data1[x2][y][2]==b[6]:
                            r7c.append(d14)
                        if data1[x2][y][2]==b[7]:
                            r8c.append(d14)
                        if data1[x2][y][2]==b[8]:
                            r9c.append(d14)
                        if data1[x2][y][2]==b[9]:
                            r10c.append(d14)
                        if data1[x2][y][2]==b[10]:
                            r11c.append(d14)
                        if data1[x2][y][2]==b[11]:
                            r12c.append(d14)
                        if data1[x2][y][2]==b[12]:
                            r13c.append(d14)
                    except:
                        pass
                
        drawing1 = Drawing(400, 200)

        if len(b) == 1:
            data2 = [
            ((sum(r1)/len(r1))),
            ((sum(r1a)/len(r1a)))
            ]
        if len(b) == 2:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))))
            ]
        if len(b) == 3:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))))
            ]
        if len(b) == 4:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))))
            ]
        if len(b) == 5:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))))
            ]
        if len(b) == 6:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))))
            ]
        if len(b) == 7:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))))
            ]
        if len(b) == 8:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))))
            ]
        if len(b) == 9:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8))), ((sum(r9)/len(r9)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))), ((sum(r9a)/len(r9a))))
            ]
        if len(b) == 10:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8))), ((sum(r9)/len(r9))), ((sum(r10)/len(r10)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))), ((sum(r9a)/len(r9a))), ((sum(r10a)/len(r10a))))
            ]
        if len(b) == 11:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8))), ((sum(r9)/len(r9))), ((sum(r10)/len(r10))), ((sum(r11)/len(r11)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))), ((sum(r9a)/len(r9a))), ((sum(r10a)/len(r10a))), ((sum(r11a)/len(r11a))))
            ]
        if len(b) == 12:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8))), ((sum(r9)/len(r9))), ((sum(r10)/len(r10))), ((sum(r11)/len(r11))), ((sum(r12)/len(r12)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))), ((sum(r9a)/len(r9a))), ((sum(r10a)/len(r10a))), ((sum(r11a)/len(r11a))), ((sum(r12a)/len(r12a))))
            ]
        if len(b) == 13:
            data2 = [
            ((sum(r1)/len(r1)), ((sum(r2)/len(r2))), ((sum(r3)/len(r3))), ((sum(r4)/len(r4))), ((sum(r5)/len(r5))), ((sum(r6)/len(r6))), ((sum(r7)/len(r7))), ((sum(r8)/len(r8))), ((sum(r9)/len(r9))), ((sum(r10)/len(r10))), ((sum(r11)/len(r11))), ((sum(r12)/len(r12))), ((sum(r13)/len(r13)))),
            ((sum(r1a)/len(r1a)), ((sum(r2a)/len(r2a))), ((sum(r3a)/len(r3a))), ((sum(r4a)/len(r4a))), ((sum(r5a)/len(r5a))), ((sum(r6a)/len(r6a))), ((sum(r7a)/len(r7a))), ((sum(r8a)/len(r8a))), ((sum(r9a)/len(r9a))), ((sum(r10a)/len(r10a))), ((sum(r11a)/len(r11a))), ((sum(r12a)/len(r12a))), ((sum(r13a)/len(r13a))))
            ]
        
        
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.height = 150
        bc.width = 400
        bc.data = data2
        bc.strokeColor = colors.black
        bc.bars[(0, 0)].fillColor = colors.lightblue
        bc.bars[(0, 1)].fillColor = colors.lightblue
        bc.bars[(0, 2)].fillColor = colors.lightblue
        bc.bars[(0, 3)].fillColor = colors.lightblue
        bc.bars[(0, 4)].fillColor = colors.lightblue
        bc.bars[(0, 5)].fillColor = colors.lightblue
        bc.bars[(0, 6)].fillColor = colors.lightblue
        bc.bars[(1, 0)].fillColor = colors.lemonchiffon
        bc.bars[(1, 1)].fillColor = colors.lemonchiffon
        bc.bars[(1, 2)].fillColor = colors.lemonchiffon
        bc.bars[(1, 3)].fillColor = colors.lemonchiffon
        bc.bars[(1, 4)].fillColor = colors.lemonchiffon
        bc.bars[(1, 5)].fillColor = colors.lemonchiffon
        bc.bars[(1, 6)].fillColor = colors.lemonchiffon
        bc.groupSpacing = 10
        bc.barSpacing = 1.5
        bc.valueAxis.valueMin = 100
        bc.valueAxis.valueMax = 115
        bc.valueAxis.valueStep = 1
        bc.barLabelFormat = '%.2f'
        bc.barLabels.dy = -20
        bc.barLabels.dx = -2
        bc.barLabels.angle = 90
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.categoryAxis.labels.dx = 7
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.labels.angle = 0
        bc.categoryAxis.categoryNames = b
        drawing1.add(bc)
        drawing1.save()
        x, y = 65, 280 # coordinates (from left bottom)
        renderPDF.draw(drawing1, pdf, x, y, showBoundary=False)
        drawing2 = Drawing(400, 200)
        if len(b) == 1:
            data2 = [
            ((sum(r1b)/len(r1b))),
            ((sum(r1c)/len(r1c)))
            ]
        if len(b) == 2:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))))
            ]
        if len(b) == 3:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))))
            ]
        if len(b) == 4:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))))
            ]
        if len(b) == 5:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))))
            ]
        if len(b) == 6:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))))
            ]
        if len(b) == 7:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))))
            ]
        if len(b) == 8:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))))
            ]
        if len(b) == 9:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b))), ((sum(r9b)/len(r9b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))), ((sum(r9c)/len(r9c))))
            ]
        if len(b) == 10:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b))), ((sum(r9b)/len(r9b))), ((sum(r10b)/len(r10b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))), ((sum(r9c)/len(r9c))), ((sum(r10c)/len(r10c))))
            ]
        if len(b) == 11:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b))), ((sum(r9b)/len(r9b))), ((sum(r10b)/len(r10b))), ((sum(r11b)/len(r11b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))), ((sum(r9c)/len(r9c))), ((sum(r10c)/len(r10c))), ((sum(r11c)/len(r11c))))
            ]
        if len(b) == 12:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b))), ((sum(r9b)/len(r9b))), ((sum(r10b)/len(r10b))), ((sum(r11b)/len(r11b))), ((sum(r12b)/len(r12b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))), ((sum(r9c)/len(r9c))), ((sum(r10c)/len(r10c))), ((sum(r11c)/len(r11c))), ((sum(r12c)/len(r12c))))
            ]
        if len(b) == 13:
            data3 = [
            ((sum(r1b)/len(r1b)), ((sum(r2b)/len(r2b))), ((sum(r3b)/len(r3b))), ((sum(r4b)/len(r4b))), ((sum(r5b)/len(r5b))), ((sum(r6b)/len(r6b))), ((sum(r7b)/len(r7b))), ((sum(r8b)/len(r8b))), ((sum(r9b)/len(r9b))), ((sum(r10b)/len(r10b))), ((sum(r11b)/len(r11b))), ((sum(r12b)/len(r12b))), ((sum(r13b)/len(r13b)))),
            ((sum(r1c)/len(r1c)), ((sum(r2c)/len(r2c))), ((sum(r3c)/len(r3c))), ((sum(r4c)/len(r4c))), ((sum(r5c)/len(r5c))), ((sum(r6c)/len(r6c))), ((sum(r7c)/len(r7c))), ((sum(r8c)/len(r8c))), ((sum(r9c)/len(r9c))), ((sum(r10c)/len(r10c))), ((sum(r11c)/len(r11c))), ((sum(r12c)/len(r12c))), ((sum(r13c)/len(r13c))))
            ]
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.height = 150
        bc.width = 400
        bc.data = data3
        bc.strokeColor = colors.black
        
        bc.bars[(0, 0)].fillColor = colors.lightblue
        bc.bars[(0, 1)].fillColor = colors.lightblue
        bc.bars[(0, 2)].fillColor = colors.lightblue
        bc.bars[(0, 3)].fillColor = colors.lightblue
        bc.bars[(0, 4)].fillColor = colors.lightblue
        bc.bars[(0, 5)].fillColor = colors.lightblue
        bc.bars[(0, 6)].fillColor = colors.lightblue
        bc.bars[(1, 0)].fillColor = colors.lemonchiffon
        bc.bars[(1, 1)].fillColor = colors.lemonchiffon
        bc.bars[(1, 2)].fillColor = colors.lemonchiffon
        bc.bars[(1, 3)].fillColor = colors.lemonchiffon
        bc.bars[(1, 4)].fillColor = colors.lemonchiffon
        bc.bars[(1, 5)].fillColor = colors.lemonchiffon
        bc.bars[(1, 6)].fillColor = colors.lemonchiffon
        bc.groupSpacing = 10
        bc.barSpacing = 1.5
        bc.valueAxis.valueMin = 100
        bc.valueAxis.valueMax = 115
        bc.valueAxis.valueStep = 1
        bc.barLabelFormat = '%.2f'
        bc.barLabels.dy = -20
        bc.barLabels.dx = -2
        bc.barLabels.angle = 90
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.categoryAxis.labels.dx = 7
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.labels.angle = 0
        bc.categoryAxis.categoryNames = b
        drawing2.add(bc)
        drawing2.save()
        x, y = 65, 80 # coordinates (from left bottom)
        renderPDF.draw(drawing2, pdf, x, y, showBoundary=False)
        
        xxx1 = -1
        datacomp = [["Class","Reading Score\n2020","Reading Score\n2021","Listening Score\n2020","Listening Score\n2021","Number of\nStudents"]]
        datacompa = [["Class","Reading Score\n2020","Reading Score\n2021","Listening Score\n2020","Listening Score\n2021","Number of\nStudents"]]
        while xxx1<len(b)-1:
            xxx1+=1
            datacomp1 = [b[xxx1],"{:.2f}".format(data2[0][xxx1]),"{:.2f}".format(data2[1][xxx1]),"{:.2f}".format(data3[0][xxx1]),"{:.2f}".format(data3[1][xxx1]),kk1[xxx1]]
            datacomp1a = [b[xxx1],data2[0][xxx1],data2[1][xxx1],data3[0][xxx1],data3[1][xxx1],kk1[xxx1]]
            datacomp.append(datacomp1)
            datacompa.append(datacomp1a)
        xxx2 = 0
        tt1,tt2,tt3,tt4,tt5=[],[],[],[],[]
        while xxx2<len(b):
            xxx2+=1
            t1 = float(datacompa[xxx2][1])
            t2 = float(datacompa[xxx2][2])
            t3 = float(datacompa[xxx2][3])
            t4 = float(datacompa[xxx2][4])
            t5 = int(datacompa[xxx2][5])
            tt1.append(t1)
            tt2.append(t2)
            tt3.append(t3)
            tt4.append(t4)
            tt5.append(t5)
        datacomp2 = ["School","{:.2f}".format(sum(tt1)/len(b)),"{:.2f}".format(sum(tt2)/len(b)),"{:.2f}".format(sum(tt3)/len(b)),"{:.2f}".format(sum(tt4)/len(b)),sum(tt5)]
        datacomp.append(datacomp2)
        table1 = Table(datacomp)
        
        style1 = TableStyle([
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('RIGHTPADDING',(0,0),(-1,-1),12),
            ('LEFTPADDING',(0,0),(-1,-1),12),
            ('BACKGROUND',(1,0),(1,0),colors.lightblue),
            ('BACKGROUND',(2,0),(2,0),colors.lemonchiffon),
            ('BACKGROUND',(3,0),(3,0),colors.lightblue),
            ('BACKGROUND',(4,0),(4,0),colors.lemonchiffon),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,-1), 8)])
        table1.setStyle(style1)
        table1.wrapOn(pdf, 150, 150)
        table1.drawOn(pdf, 60,550)
        pdf.showPage()
                
    ay1 = str(sheet1['H3'].value)[5:-12]
    ay2 = str(sheet2['H3'].value)[5:-12]
    yil1 = str(sheet1['H3'].value)[:4]
    yil2 = str(sheet2['H3'].value)[:4]
    if ay1 > "00":
        if ay1 == "01":
            ay1 = "Jan"
        elif ay1 == "02":
            ay1 = "Feb"
        elif ay1 == "03":
            ay1 = "Mar"
        elif ay1 == "04":
            ay1 = "Apr"
        elif ay1 == "05":
            ay1 = "May" 
        elif ay1 == "06":
            ay1 = "Jun"
        elif ay1 == "07":
            ay1 = "Jul"
        elif ay1 == "08":
            ay1 = "Aug"
        elif ay1 == "09":
            ay1 = "Sep"
        elif ay1 == "10":
            ay1 = "Oct"
        elif ay1 == "11":
            ay1 = "Nov"
        elif ay1 == "12":
            ay1 = "Dec"
    if ay2 > "00":
        if ay2 == "01":
            ay2 = "Jan"
        elif ay2 == "02":
            ay2 = "Feb"
        elif ay2 == "03":
            ay2 = "Mar"
        elif ay2 == "04":
            ay2 = "Apr"
        elif ay2 == "05":
            ay2 = "May" 
        elif ay2 == "06":
            ay2 = "Jun"
        elif ay2 == "07":
            ay2 = "Jul"
        elif ay2 == "08":
            ay2 = "Aug"
        elif ay2 == "09":
            ay2 = "Sep"
        elif ay2 == "10":
            ay2 = "Oct"
        elif ay2 == "11":
            ay2 = "Nov"
        elif ay2 == "12":
            ay2 = "Dec"
    okuladi = str(sheet1['A3'].value)
    fileName = okuladi+' - Reading and Listening Report.pdf'
    outf = outputfolder+"/"+fileName
    pdf = canvas.Canvas(outputfolder+"/"+fileName, pagesize=pagesizes.letter)
    pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
    pdf.setFont('Helvetica',18)
    pdf.drawCentredString(307.5,480,okuladi)
    pdf.drawCentredString(307.5,430,'COMPARATIVE REPORT')
    pdf.drawCentredString(307.5,380,str(ay1).upper()+" "+str(yil1)+" - "+str(ay2).upper()+" "+str(yil2))
    pdf.showPage()

    for x1 in range(len(b)):
        sonuclar()
        pdf.showPage()
    chart1()

    pdf.save()
    ############## Grafik Sayfasını 2.Sayfa olarak ayarlayan bölüm
    import os,glob
    from PyPDF2 import PdfFileReader, PdfFileWriter,PdfFileMerger
    path = outf
    pdf = PdfFileReader(path)
    for page in range(pdf.getNumPages()):
        pdf_writer = PdfFileWriter()
        pdf_writer.addPage(pdf.getPage(page))

        output_filename = '{}_{}XXXSamet.pdf'.format("f",page+1)

        with open(output_filename, 'wb') as out:
            pdf_writer.write(out)

    pdfs1 = glob.glob("*XXXSamet.pdf")
    os.rename("f_1XXXSamet.pdf","f_0XXXSamet.pdf")
    os.rename(pdfs1[-1],"f_1XXXSamet.pdf")
    pdfs2 = glob.glob("*XXXSamet.pdf")

    merger = PdfFileMerger()

    for pdf1 in pdfs2:
        merger.append(pdf1)

    merger.write(outf)
    merger.close()
    for pdf2 in pdfs2:
        os.remove(pdf2)
    #########Bu bölümün bitişi

def rlsreport():
    book1 = openpyxl.load_workbook(sinav1)
    sheet1 = book1['Sheet2']
    sheetlenght1 = len(sheet1['A'])
    book2 = openpyxl.load_workbook(sinav2)
    sheet2 = book2['Sheet2']
    book3 = openpyxl.load_workbook(sinav3)
    sheet3 = book3['Sheet2']
    classlist = []
    for row in sheet1.rows:
        classlist.append(row[2].value)
    my_dict = {i:classlist.count(i) for i in classlist}
    b = list(set([x for x in classlist if classlist.count(x) > 2]))
    b.sort()
    filigram = './filigram.jpg'
    data1= []
    def sonuclar():
        satir = 2
        toplamsatir = int(sheetlenght1)
        chartdata = []
        data= [['','Student Name', 'Class', 'Reading\nScore\nCEFR\n'+str(ay1)+'\n'+str(yil1),'', 'Reading\nScore\nCEFR\n'+str(ay2)+'\n'+str(yil2),'', 'Listening\nScore\nCEFR\n'+str(ay1)+'\n'+str(yil1),'','Listening\nScore\nCEFR\n'+str(ay2)+'\n'+str(yil2),'','Speaking\nScore\nCEFR\n'+str(ay3)+'\n'+str(yil3),'']]
        while satir<toplamsatir:
            satir +=1
            studentnumber = int(sheet1['F'+str(satir)].value)
            studentname = str(sheet1['E'+str(satir)].value)+str(" ")+str(sheet1['D'+str(satir)].value)
            sinif = sheet1['C'+str(satir)].value
            rs1a = sheet1['I'+str(satir)].value
            rs1b = sheet1['J'+str(satir)].value
            ls1a = sheet1['M'+str(satir)].value
            ls1b = sheet1['N'+str(satir)].value
            for row in sheet2.rows:
                for any_cell in row:
                    if any_cell.value == studentnumber:
                            rs2a = sheet2.cell(row=any_cell.row, column=9).value
                            rs2b = sheet2.cell(row=any_cell.row, column=10).value
                            ls2a = sheet2.cell(row=any_cell.row, column=13).value
                            ls2b = sheet2.cell(row=any_cell.row, column=14).value
            for row in sheet3.rows:
                for any_cell in row:
                    if any_cell.value == studentnumber:
                            ss1a = sheet3.cell(row=any_cell.row, column=9).value
                            ss1b = sheet3.cell(row=any_cell.row, column=10).value
            if b[x1] == sinif:
                rs1b = rs1b.replace('Below A1','*')
                ls1b = ls1b.replace('Below A1','*')
                rs2b = rs2b.replace('Below A1','*')
                ls2b = ls2b.replace('Below A1','*')
                ss1b = ss1b.replace('Below A1','*')
                tumveriler = ["",studentname, sinif, rs1a, rs1b, ls1a,ls1b,rs2a,rs2b,ls2a,ls2b,ss1a,ss1b]
                chartdata.append(tumveriler)
                data.append(tumveriler)
                table = Table(data)
                style = TableStyle([
                    ('GRID',(1,0),(-1,-1),0.5,colors.grey),
                    ('GRID',(1,1),(-1,-1),0.5,colors.grey),
                    ('ALIGN',(0,0),(-1,0),'CENTER'),
                    ('VALIGN',(0,0),(-1,0),'MIDDLE'),
                    ('TOPPADDING',(0,0),(-1,0),12),
                    ('BOTTOMPADDING',(0,0),(-1,0),12),
                    ('RIGHTPADDING',(1,0),(1,0),38),
                    ('LEFTPADDING',(1,0),(1,0),38),
                    
                    ('ALIGN',(2,0),(-1,-1),'CENTER'),
                    ('VALIGN',(2,0),(-1,1),'MIDDLE'),
                    ('RIGHTPADDING',(3,0),(-1,0),10),
                    ('LEFTPADDING',(3,0),(-1,0),10),
                    
                    ('ALIGN',(0,1),(0,-1),'CENTER'),
                    ('VALIGN',(0,1),(0,-1),'MIDDLE'),
                    
                    ('BACKGROUND',(3,0),(4,0),colors.lightblue),
                    ('BACKGROUND',(5,0),(6,0),colors.lemonchiffon),
                    ('BACKGROUND',(7,0),(8,0),colors.lightblue),
                    ('BACKGROUND',(9,0),(10,0),colors.lemonchiffon),
                    ('BACKGROUND',(11,0),(12,0),colors.lemonchiffon),
                    ('SPAN',(3,0),(4,0)),
                    ('SPAN',(5,0),(6,0)),
                    ('SPAN',(7,0),(8,0)),
                    ('SPAN',(9,0),(10,0)),
                    ('SPAN',(11,0),(12,0)),
                    ('FONTSIZE', (0,1), (-1,-1), 8)
                ])
                u1 = 680
                k1 = my_dict[b[x1]]
                pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
                pdf.setFont('Helvetica',10)
                pdf.drawString(72,u1,'Class : '+str(b[x1]))
                pdf.drawString(170,u1,'Total Student Number : '+str(k1))
                pdf.setFont('Helvetica',9)
                pdf.drawString(75,100,'CEFR : The Common European Framework of Reference')
                pdf.drawString(75,87,'* : Below A1')
                pdf.drawString(75,74,'- : Do not take the test')
                table.setStyle(style)
                table.wrapOn(pdf, 1500, 50)
                table.drawOn(pdf, 60,(u1-95-18*k1))
        data1.append(chartdata)
    def chart1():
        pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
        
        pdf.drawCentredString(307.5,625,"Reading Score Comparison 2020-2021")
        pdf.drawCentredString(307.5,365,"Listening Score Comparison 2020-2021")
        pdf.setFont('Helvetica',9)
        pdf.drawString(95, 123,"2020 :")
        pdf.drawString(95, 103,"2021 :")
        
        drawing3 = Drawing(400, 200)
        drawing3.add(Rect(10, 10, 70, 13, fillColor=colors.lightblue))
        x, y = 120, 110 # coordinates (from left bottom)
        renderPDF.draw(drawing3, pdf, x, y, showBoundary=False)
        drawing4 = Drawing(400, 200)
        drawing4.add(Rect(10, 10, 70, 13, fillColor=colors.lemonchiffon))
        x, y = 120, 90 # coordinates (from left bottom)
        renderPDF.draw(drawing4, pdf, x, y, showBoundary=False)

        x2=-1
        r1,r2,r3,r4,r5,r6,r7,r8,r9,r1a,r1b,r1c,r2a,r2b,r2c,r3a,r3b,r3c,r4a,r4b,r4c,r5a,r5b,r5c,r6a,r6b,r6c,r7a,r7b,r7c,r8a,r8b,r8c,r9a,r9b,r9c=[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
        while x2<len(data1)-1:
            x2+=1
            y=-1
            while y<len(data1[x2])-1:
                y+=1
                d11 = data1[x2][y][3]
                d12 = data1[x2][y][5]
                d13 = data1[x2][y][7]
                d14 = data1[x2][y][9]
                try:
                    if data1[x2][y][2]==b[0]:
                        r1.append(d11)
                        r1a.append(d12)
                        r1b.append(d13)
                        r1c.append(d14)
                    if data1[x2][y][2]==b[1]:
                        r2.append(d11)
                        r2a.append(d12)
                        r2b.append(d13)
                        r2c.append(d14)
                    if data1[x2][y][2]==b[2]:
                        r3.append(d11)
                        r3a.append(d12)
                        r3b.append(d13)
                        r3c.append(d14)
                    if data1[x2][y][2]==b[3]:
                        r4.append(d11)
                        r4a.append(d12)
                        r4b.append(d13)
                        r4c.append(d14)
                    if data1[x2][y][2]==b[4]:
                        r5.append(d11)
                        r5a.append(d12)
                        r5b.append(d13)
                        r5c.append(d14)
                    if data1[x2][y][2]==b[5]:
                        r6.append(d11)
                        r6a.append(d12)
                        r6b.append(d13)
                        r6c.append(d14)
                    if data1[x2][y][2]==b[6]:
                        r7.append(d11)
                        r7a.append(d12)
                        r7b.append(d13)
                        r7c.append(d14)
                    if data1[x2][y][2]==b[7]:
                        r8.append(d11)
                        r8a.append(d12)
                        r8b.append(d13)
                        r8c.append(d14)
                    if data1[x2][y][2]==b[8]:
                        r9.append(d11)
                        r9a.append(d12)
                        r9b.append(d13)
                        r9c.append(d14)
                except:
                    pass
        drawing1 = Drawing(400, 200)
        data2 = [
        (int(sum(r1)/len(r1)), (int(sum(r2)/len(r2))), (int(sum(r3)/len(r3))), (int(sum(r4)/len(r4))), (int(sum(r5)/len(r5))), (int(sum(r6)/len(r6)))),
        (int(sum(r1a)/len(r1a)), (int(sum(r2a)/len(r2a))), (int(sum(r3a)/len(r3a))), (int(sum(r4a)/len(r4a))), (int(sum(r5a)/len(r5a))), (int(sum(r6a)/len(r6a))))
        ]
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.height = 175
        bc.width = 400
        bc.data = data2
        bc.strokeColor = colors.black
        bc.bars[(0, 0)].fillColor = colors.lightblue
        bc.bars[(0, 1)].fillColor = colors.lightblue
        bc.bars[(0, 2)].fillColor = colors.lightblue
        bc.bars[(0, 3)].fillColor = colors.lightblue
        bc.bars[(0, 4)].fillColor = colors.lightblue
        bc.bars[(0, 5)].fillColor = colors.lightblue
        bc.bars[(0, 6)].fillColor = colors.lightblue
        bc.bars[(1, 0)].fillColor = colors.lemonchiffon
        bc.bars[(1, 1)].fillColor = colors.lemonchiffon
        bc.bars[(1, 2)].fillColor = colors.lemonchiffon
        bc.bars[(1, 3)].fillColor = colors.lemonchiffon
        bc.bars[(1, 4)].fillColor = colors.lemonchiffon
        bc.bars[(1, 5)].fillColor = colors.lemonchiffon
        bc.bars[(1, 6)].fillColor = colors.lemonchiffon
        bc.groupSpacing = 10
        bc.barSpacing = 1.5
        bc.valueAxis.valueMin = 100
        bc.valueAxis.valueMax = 115
        bc.valueAxis.valueStep = 1
        bc.barLabelFormat = '%.0f'
        bc.barLabels.dy = 8
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.categoryAxis.labels.dx = 7
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.labels.angle = 0
        bc.categoryAxis.categoryNames = b
        drawing1.add(bc)
        drawing1.save()
        x, y = 65, 380 # coordinates (from left bottom)
        renderPDF.draw(drawing1, pdf, x, y, showBoundary=False)
        drawing2 = Drawing(400, 200)
        data3 = [
        (int(sum(r1b)/len(r1b)), (int(sum(r2b)/len(r2b))), (int(sum(r3b)/len(r3b))), (int(sum(r4b)/len(r4b))), (int(sum(r5b)/len(r5b))), (int(sum(r6b)/len(r6b)))),
        (int(sum(r1c)/len(r1c)), (int(sum(r2c)/len(r2c))), (int(sum(r3c)/len(r3c))), (int(sum(r4c)/len(r4c))), (int(sum(r5c)/len(r5c))), (int(sum(r6c)/len(r6c))))
        ]
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.height = 175
        bc.width = 400
        bc.data = data3
        bc.strokeColor = colors.black
        
        bc.bars[(0, 0)].fillColor = colors.lightblue
        bc.bars[(0, 1)].fillColor = colors.lightblue
        bc.bars[(0, 2)].fillColor = colors.lightblue
        bc.bars[(0, 3)].fillColor = colors.lightblue
        bc.bars[(0, 4)].fillColor = colors.lightblue
        bc.bars[(0, 5)].fillColor = colors.lightblue
        bc.bars[(0, 6)].fillColor = colors.lightblue
        bc.bars[(1, 0)].fillColor = colors.lemonchiffon
        bc.bars[(1, 1)].fillColor = colors.lemonchiffon
        bc.bars[(1, 2)].fillColor = colors.lemonchiffon
        bc.bars[(1, 3)].fillColor = colors.lemonchiffon
        bc.bars[(1, 4)].fillColor = colors.lemonchiffon
        bc.bars[(1, 5)].fillColor = colors.lemonchiffon
        bc.bars[(1, 6)].fillColor = colors.lemonchiffon
        bc.groupSpacing = 10
        bc.barSpacing = 1.5
        bc.valueAxis.valueMin = 100
        bc.valueAxis.valueMax = 115
        bc.valueAxis.valueStep = 1
        bc.barLabelFormat = '%.0f'
        bc.barLabels.dy = 8
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.categoryAxis.labels.dx = 7
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.labels.angle = 0
        bc.categoryAxis.categoryNames = b
        drawing2.add(bc)
        drawing2.save()
        x, y = 65, 120 # coordinates (from left bottom)
        renderPDF.draw(drawing2, pdf, x, y, showBoundary=False)
        pdf.showPage()

    ay1 = str(sheet1['H3'].value)[5:-12]
    ay2 = str(sheet2['H3'].value)[5:-12]
    ay3 = str(sheet3['H3'].value)[5:-12]
    yil1 = str(sheet1['H3'].value)[:4]
    yil2 = str(sheet2['H3'].value)[:4]
    yil3 = str(sheet3['H3'].value)[:4]
    if ay1 > "00":
        if ay1 == "01":
            ay1 = "Jan"
        elif ay1 == "02":
            ay1 = "Feb"
        elif ay1 == "03":
            ay1 = "Mar"
        elif ay1 == "04":
            ay1 = "Apr"
        elif ay1 == "05":
            ay1 = "May" 
        elif ay1 == "06":
            ay1 = "Jun"
        elif ay1 == "07":
            ay1 = "Jul"
        elif ay1 == "08":
            ay1 = "Aug"
        elif ay1 == "09":
            ay1 = "Sep"
        elif ay1 == "10":
            ay1 = "Oct"
        elif ay1 == "11":
            ay1 = "Nov"
        elif ay1 == "12":
            ay1 = "Dec"
    if ay2 > "00":
        if ay2 == "01":
            ay2 = "Jan"
        elif ay2 == "02":
            ay2 = "Feb"
        elif ay2 == "03":
            ay2 = "Mar"
        elif ay2 == "04":
            ay2 = "Apr"
        elif ay2 == "05":
            ay2 = "May" 
        elif ay2 == "06":
            ay2 = "Jun"
        elif ay2 == "07":
            ay2 = "Jul"
        elif ay2 == "08":
            ay2 = "Aug"
        elif ay2 == "09":
            ay2 = "Sep"
        elif ay2 == "10":
            ay2 = "Oct"
        elif ay2 == "11":
            ay2 = "Nov"
        elif ay2 == "12":
            ay2 = "Dec"
    if ay3 > "00":
        if ay3 == "01":
            ay3 = "Jan"
        elif ay3 == "02":
            ay3 = "Feb"
        elif ay3 == "03":
            ay3 = "Mar"
        elif ay3 == "04":
            ay3 = "Apr"
        elif ay3 == "05":
            ay3 = "May" 
        elif ay3 == "06":
            ay3 = "Jun"
        elif ay3 == "07":
            ay3 = "Jul"
        elif ay3 == "08":
            ay3 = "Aug"
        elif ay3 == "09":
            ay3 = "Sep"
        elif ay3 == "10":
            ay3 = "Oct"
        elif ay3 == "11":
            ay3 = "Nov"
        elif ay3 == "12":
            ay3 = "Dec"

    okuladi = str(sheet1['A3'].value)
    fileName = okuladi+' - Reading and Listening - Speaking Report.pdf'
    pdf = canvas.Canvas(fileName, pagesize=pagesizes.letter)
    pdf.drawImage(filigram, 0,-3, width=615,height=790,mask=None)
    pdf.setFont('Helvetica',18)
    pdf.drawCentredString(307.5,480,okuladi)
    pdf.drawCentredString(307.5,430,'COMPARATIVE REPORT')
    pdf.drawCentredString(307.5,380,str(ay1).upper()+" "+str(yil1)+" - "+str(ay2).upper()+" "+str(yil2))
    pdf.showPage()
    for x1 in range(len(b)):
        sonuclar()
        pdf.showPage()
    chart1()
    pdf.save()

def rlreportbutton():
    tr1 = threading.Thread(target=rlreport)
    tr1.start()
    
def rlsreportbutton():
    tr2 = threading.Thread(target=rlsreport)
    tr2.start()